import streamlit as st
import pandas as pd
import math
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Font
from openpyxl.utils import get_column_letter
import os, json
from datetime import datetime

# ---------------- CONFIG ----------------
DENSITY = 7850
SAVE_DIR = "saved_estimations"
os.makedirs(SAVE_DIR, exist_ok=True)

st.set_page_config(page_title="Site Estimation", layout="wide")
st.title("SITE ESTIMATION SHEET")

# ======================================================
# LOAD SAVED ESTIMATION
# ======================================================
st.sidebar.title("REFERENCE ESTIMATIONS")
files = [""] + sorted(os.listdir(SAVE_DIR))
selected = st.sidebar.selectbox("VIEW SAVED ESTIMATION", files)

loaded = None
if selected:
    with open(os.path.join(SAVE_DIR, selected)) as f:
        loaded = json.load(f)

# ======================================================
# MATERIAL TABLE
# ======================================================
material_items = [
    "BED BOLT","BASE PLATE","WALL PLATE","FASTENER BOLT / CHEMICAL BOLT",
    "TRIANGLE CLEAT","TRUSS CLEAT","RUNNER CLEAT",
    "ROUND POST","SQUARE POST","TRUSS ROUND","TRUSS SQUARE","TRUSS INTERVAL",
    "FRONT PURLIN","CENTER PURLIN","BACK PURLIN",
    "HANGING PIPE","STAY PIPE","BENDING PIPE CHARGES",
    "HOSPITAL TRACK HEAVY","ONE WAY / SINGLE WAY SECTION",
    "TWO WAY / DOUBLE WAY SECTION","4MM CORNER / MIDDLE PLATE M.S SET",
    "PVC INSULATED 8MM STEEL ROPE","6 MM BOLT NUT WASHER",
    "SS THREADED BOLT (4\" HOLES ROD)","KEDER (12 MM / 9 MM)",
    "SELF DRILLING SCREW","\"U\" CLAMP","D SHACKLES",
    "12MM SS ROD – 6\"","12MM S.S. NUT","END CAP NUT"
]

n = len(material_items)
extra = n - 19

material_df = pd.DataFrame({
    "ITEM": material_items,
    "DIAMETER (MM)": [0,None,None,None,None,None,None,0,None,0,None,0,0,0,0,0,0,None,None] + [None]*extra,
    "LENGTH (MM)": [0,None,None,None,None,None,None,0,0,0,0,0,0,0,0,0,0,None,None] + [None]*extra,
    "SIDE A (MM)": [None,0,0,0,0,0,0,None,0,None,0,0,0,0,0,0,0,None,None] + [None]*extra,
    "SIDE B (MM)": [None,0,0,0,0,0,0,None,0,None,0,0,0,0,0,0,0,None,None] + [None]*extra,
    "THICKNESS (MM)": [None,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,None,None] + [None]*extra,
    "PIPE": [""]*19 + [None]*extra,
    "SQFT": [0]*19 + [None]*extra,
    "QTY": [0]*n,
    "RATE (₹)": [0.0]*n
})

if loaded:
    material_df = pd.DataFrame(loaded["material"])

material_df = st.data_editor(material_df, disabled=["ITEM"], use_container_width=True)

# ======================================================
# MATERIAL CALCULATIONS
# ======================================================
def round_bar(d,l,q):
    return math.pi*(d/2000)**2*(l/1000)*DENSITY*q if d and l and q else 0

def plate(a,b,t,q):
    return (a*b*t/1e9)*DENSITY*q if a and b and t and q else 0

def round_tube(d,t,l,q):
    if not all([d,t,l,q]): return 0
    D,T,L=d/1000,t/1000,l/1000
    di=D-2*T
    return math.pi*((D/2)**2-(di/2)**2)*L*DENSITY*q if di>0 else 0

def square_tube(a,b,t,l,q):
    if not all([a,b,t,l,q]): return 0
    A,B,T,L=a/1000,b/1000,t/1000,l/1000
    Ai,Bi=A-2*T,B-2*T
    return (A*B-Ai*Bi)*L*DENSITY*q if Ai>0 and Bi>0 else 0

material_df["_WEIGHT"] = 0.0
material_df["_VALUE"] = 0.0
qty_rate_only_items = material_items[18:]

for i,row in material_df.iterrows():
    q,r,sqft = row["QTY"], row["RATE (₹)"], row["SQFT"]
    wt = 0

    if row["ITEM"] == "BENDING PIPE CHARGES":
        material_df.at[i,"_VALUE"] = sqft * r
        continue

    if row["ITEM"] in qty_rate_only_items:
        material_df.at[i,"_VALUE"] = q * r
        continue

    if row["ITEM"]=="BED BOLT":
        wt=round_bar(row["DIAMETER (MM)"],row["LENGTH (MM)"],q)
    elif row["ITEM"] in ["BASE PLATE","WALL PLATE","TRIANGLE CLEAT","TRUSS CLEAT","RUNNER CLEAT"]:
        wt=plate(row["SIDE A (MM)"],row["SIDE B (MM)"],row["THICKNESS (MM)"],q)
    elif row["ITEM"] in ["ROUND POST","TRUSS ROUND"]:
        wt=round_tube(row["DIAMETER (MM)"],row["THICKNESS (MM)"],row["LENGTH (MM)"],q)
    elif row["ITEM"] in ["SQUARE POST","TRUSS SQUARE"]:
        wt=square_tube(row["SIDE A (MM)"],row["SIDE B (MM)"],row["THICKNESS (MM)"],row["LENGTH (MM)"],q)

    material_df.at[i,"_WEIGHT"] = wt
    material_df.at[i,"_VALUE"] = wt * r

# ======================================================
# LABOUR TABLE
# ======================================================
st.title("LABOUR CHARGES")

labour_df = pd.DataFrame({
    "ITEM":["SUPERVISOR","WELDER","PAINTER","HELPER"],
    "QTY":[0]*4,
    "DAYS":[0]*4,
    "RATE (₹)":[0.0]*4
})

if loaded:
    labour_df = pd.DataFrame(loaded["labour"])

labour_df = st.data_editor(labour_df, disabled=["ITEM"], use_container_width=True)

# ======================================================
# FABRIC TABLE
# ======================================================
st.title("FABRIC CALCULATION")

fabric_df = pd.DataFrame({
    "ITEM":["600 GSM","650 GSM","700 GSM","730 GSM","900 GSM"],
    "FABRIC":["LUCKY"]*5,
    "SQFT":[0]*5,
    "RATE (₹)":[0.0]*5
})

if loaded:
    fabric_df = pd.DataFrame(loaded["fabric"])

fabric_df = st.data_editor(
    fabric_df,
    disabled=["ITEM"],
    column_config={
        "FABRIC": st.column_config.SelectboxColumn(
            "FABRIC",
            options=["CSC","LUCKY","KHOSLA"]
        )
    },
    use_container_width=True
)

# ======================================================
# FINAL ESTIMATION (NO DUPLICATES)
# ======================================================
st.title("FINAL ESTIMATION")

material_final = material_df.drop(columns=["_WEIGHT","_VALUE"]).assign(
    DAYS=None,
    FABRIC=None,
    **{
        "TOTAL WEIGHT (KG)": material_df["_WEIGHT"].round(2),
        "TOTAL VALUE (₹)": material_df["_VALUE"].round(2)
    }
)

labour_final = labour_df.assign(
    **{
        "TOTAL WEIGHT (KG)": None,
        "TOTAL VALUE (₹)": labour_df["QTY"] * labour_df["DAYS"] * labour_df["RATE (₹)"]
    }
)

fabric_final = fabric_df.assign(
    **{
        "TOTAL WEIGHT (KG)": None,
        "TOTAL VALUE (₹)": fabric_df["SQFT"] * fabric_df["RATE (₹)"]
    }
)

final_df = pd.concat([material_final, labour_final, fabric_final], ignore_index=True)
final_df = final_df[final_df["TOTAL VALUE (₹)"].fillna(0) > 0]

st.dataframe(final_df, use_container_width=True)

# ======================================================
# GRAND TOTALS
# ======================================================
st.subheader("GRAND TOTALS")
c1,c2,c3,c4 = st.columns(4)

c1.metric("MATERIAL TOTAL (₹)", round(material_df["_VALUE"].sum(),2))
c2.metric("LABOUR TOTAL (₹)", round((labour_df["QTY"]*labour_df["DAYS"]*labour_df["RATE (₹)"]).sum(),2))
c3.metric("GRAND TOTAL (₹)", round(final_df["TOTAL VALUE (₹)"].sum(),2))
c4.metric("TOTAL WEIGHT (KG)", round(material_df["_WEIGHT"].sum(),2))

# ======================================================
# SAVE ESTIMATION
# ======================================================
st.subheader("SAVE ESTIMATION")
name = st.text_input("ESTIMATION NAME")

if st.button("SAVE FINAL ESTIMATION") and name:
    json.dump(
        {
            "material": material_df.drop(columns=["_WEIGHT","_VALUE"]).to_dict(),
            "labour": labour_df.to_dict(),
            "fabric": fabric_df.to_dict(),
            "saved_on": datetime.now().isoformat()
        },
        open(os.path.join(SAVE_DIR,f"{name}.json"),"w")
    )
    st.success("ESTIMATION SAVED")

# ======================================================
# EXCEL EXPORT
# ======================================================
buffer = BytesIO()
with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
    final_df.to_excel(writer, sheet_name="ESTIMATION", startrow=15, index=False)

buffer.seek(0)
wb = load_workbook(buffer)
ws = wb["ESTIMATION"]

bold = Font(bold=True)
thick = Border(*(Side(style="medium"),)*4)

ws["A2"]="SITE :-"; ws["D2"]="REP :-"; ws["H2"]="DATE :-"
for c in ["A2","D2","H2"]:
    ws[c].font = bold

for r in ws.iter_rows(min_row=16, max_row=ws.max_row):
    for c in r:
        c.border = thick

for col in range(1, ws.max_column+1):
    ws.column_dimensions[get_column_letter(col)].width = 24

final_buffer = BytesIO()
wb.save(final_buffer)
final_buffer.seek(0)

st.download_button(
    "DOWNLOAD PROFESSIONAL EXCEL",
    final_buffer,
    "FINAL_ESTIMATION.xlsx",
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)
